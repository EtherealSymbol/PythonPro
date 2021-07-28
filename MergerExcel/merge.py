import openpyxl
from extract import extract_to_dict

if __name__ == '__main__':
    # 提取表2数据
    dict_data = extract_to_dict("表2.xlsx")

    # 新建excel
    wb2 = openpyxl.Workbook()
    wb2.save('合并后.xlsx')
    print('新建成功')

    # 读取数据
    wb1 = openpyxl.load_workbook('表1.xlsx')
    wb2 = openpyxl.load_workbook('合并后.xlsx')
    sheets1 = wb1.sheetnames  # 获取sheet页
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]

    max_row = sheet1.max_row  # 最大行数
    max_column = sheet1.max_column  # 最大列数
    # 在最后一列追加一列数据：“投入”
    add_col = max_column + 1
    sheet1.insert_cols(add_col)

    # 记住产品编号数据
    pro_num = ''
    for m in range(1, max_row + 1):
        for n in range(97, 97 + max_column):  # chr(97)='a'
            n = chr(n)  # ASCII字符
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value  # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到合并后的单元格
            if n is 'c':
                pro_num = str(cell1)
        c = chr(96 + add_col)
        j = '%s%d' % (c, m)  # 单元格编号
        if m == 3:
            sheet2[j].value = "投入"
        if m > 3:
            sheet2[j].value = dict_data.get(pro_num, 0)

    wb2.save('合并后.xlsx')  # 保存数据
    print("合并完成")
    wb1.close()  # 关闭excel
    wb2.close()
    # raw_input('Press the enter key to exit.')
    input("按回车退出！")



