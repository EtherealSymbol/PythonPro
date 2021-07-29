import re
import openpyxl
from data_process import extract_num
from openpyxl.styles import colors, PatternFill

wb = openpyxl.load_workbook("西林瓶认证厂家.xlsx", data_only=True)
# wb = openpyxl.load_workbook("混悬剂.xlsx", data_only=True)
ws = wb["混悬剂GMP"]
# ws = wb["混悬剂许可证"]

# ws = wb["小容量注射剂GMP证书"]
# ws = wb["小容量注射剂许可证"]
# ws.insert_cols(2)
# ws = wb["冻干粉GMP证书"]
# print(ws)
# 参考
# ws = wb["冻干粉许可证"]
# ws = wb["小容量注射剂许可证"]
# print(ws)
# ws = wb["小容量注射剂GMP证书"]
# ws = wb["冻干粉GMP证书"]
rows = ws.rows
# print(rows)
list_rows = list(rows)
next_rows = list_rows[0:]


reg_num = r"\d+"

i = 1
# for row in next_rows:
#     for cell in row:
#         value = cell(row=i, column=1).value
#         if value is not None:
#             cap = re.search(reg, value)
#             print(cap)
#             # num = re.search(reg_num, cell.value).group()
#             # print(num)
#             print(value)
#             i+=1
list1 = []
dict = {}
# reg = r"\..+"
for rowNumber in range(1, ws.max_row + 1):
    value = ws.cell(row=rowNumber, column=2).value

    if value is not None:
        # cap = re.search(reg, value).group()
        # print(cap[1:])
        # list.append(cap[1:])
        # print(value)
        list1.append(value)
        dict[value] = 1
        # print(value)
# for l in list:
    # ws.cell(row=i, column=3).value = l
    # i+=1
# for key in dict:
    # print(key, dict.get(key))

# 改变
# ws1 = wb["混悬剂GMP"]
ws1 = wb["混悬剂许可证"]
# ws1 = wb["小容量注射剂许可证"]

# ws1 = wb["冻干粉许可证"]
# print(ws1)
# ws1 = wb["冻干粉GMP证书"]
# ws1 = wb["小容量注射剂GMP证书"]
rows1 = ws1.rows

num = 0
# print(rows1)
list_rows1 = list(rows1)
next_rows1 = list_rows1[0:]
for rowNumber1 in range(1, ws1.max_row + 1):
    value1 = ws1.cell(row=rowNumber1, column=2).value
    if value1 is not None:
        # print(value1)
        if dict.get(value1) is not None:
            num += 1
            yellow_fill = PatternFill(fill_type='solid', fgColor="FFff00")
            ws1.row_dimensions[rowNumber1].fill = yellow_fill
print(num)
# wb.save('混悬剂.xlsx')  # 保存数据
wb.save('西林瓶认证厂家.xlsx')