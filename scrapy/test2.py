import re
import openpyxl
from data_process import extract_num
from openpyxl.styles import colors, PatternFill


# wb = openpyxl.load_workbook("西林瓶认证厂家.xlsx", data_only=True)
wb = openpyxl.load_workbook("混悬剂.xlsx", data_only=True)
# ws = wb["混悬剂GMP认证"]
ws = wb["混悬剂许可证"]
# ws = wb["小容量注射剂GMP证书"]
# ws = wb["小容量注射剂许可证"]
# ws.insert_cols(2)
# ws = wb["冻干粉GMP证书"]
# ws = wb["冻干粉许可证"]
# print(ws)
rows = ws.rows
list_rows = list(rows)
next_rows = list_rows[0:]

reg = r"\.."
reg_num = r"\d+"

i=1
for row in next_rows:
    for cell in row:
        if cell.value is not None:
            # cap = re.search(reg, cell.value).group()
            # print(cap[1])
            num = re.search(reg_num, cell.value).group()
            print(num)
            print(cell.value)
            ws.cell(row=i, column=2).value=num
            i+=1


wb.save('混悬剂.xlsx')  # 保存数据





