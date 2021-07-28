import re
import openpyxl
from data_process import extract_num
from openpyxl.styles import colors, PatternFill


# wb = openpyxl.load_workbook("西林瓶认证厂家.xlsx", data_only=True)
# ws = wb["小容量注射剂GMP证书"]
# ws = wb["小容量注射剂许可证"]
# ws.insert_cols(2)
# ws = wb["冻干粉GMP证书"]
# print(ws)
# ws = wb["冻干粉许可证"]
wb = openpyxl.load_workbook("混悬剂.xlsx", data_only=True)
ws = wb["混悬剂GMP认证"]
# ws = wb["混悬剂许可证"]
rows = ws.rows
list_rows = list(rows)
next_rows = list_rows[0:]


reg_num = r"\d+"

i = 1
# for row in next_rows:
#     for cell in row:
#         value = cell(row=i, column=1).value
#         if value is not None:
#             cap = re.search(reg, value)
#             print(cap)
#             # num = re.search(reg_num, cell.value).group()
#             # print(num)
#             print(value)
#             i+=1
list = []
reg = r"\..+"
for rowNumber in range(1, ws.max_row + 1):
    value = ws.cell(row=rowNumber, column=1).value

    if value is not None:
        cap = re.search(reg, value).group()
        print(cap[1:])
        list.append(cap[1:])
        # print(value)
        # list.append(value)
for l in list:
    ws.cell(row=i, column=3).value = l
    i+=1


# wb.save('西林瓶认证厂家.xlsx')  # 保存数据
wb.save('混悬剂.xlsx')




