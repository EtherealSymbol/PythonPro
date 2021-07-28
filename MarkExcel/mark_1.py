import re
import openpyxl
from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill

if __name__ == '__main__':
    wb = openpyxl.load_workbook("表.xlsx", data_only=True)
    ws = wb.active
    # 获得所有列
    # for col in ws.columns:
    #     print(col)
    rows = ws.rows
    # 获得第一行
    # print(list(rows))
    list_rows = list(rows)
    first_row = list_rows[0]

    # 初始化物料，产品规格，数量为-1
    wl = -1
    cp = -1
    sl = -1
    k = 1
    for elem in first_row:
        # 获得“物料”所在列
        if elem.value == "物料" and wl == -1:
            wl = k
        elif elem.value == "产品规格" and cp == -1:
            cp = k
        elif elem.value == "数量" and sl == -1:
            sl = k
        else:
            k += 1
    # print(wl, cp, sl)
    next_rows = list_rows[1:]
    # 正则
    reg = r"安瓿"
    reg_ch = r"彩盒"
    reg_sms = r"说明书"
    reg_tp = r"托盘"
    i = 2
    for row in next_rows:
        for cell in row:
            wl_data = ws.cell(i, wl+1).value
            # 数量
            sl_num = ws.cell(i, sl + 2).value
            # 是安瓿
            if wl_data is not None and re.search(reg, str(wl_data)):
                # 产品规格字符串
                cp_str = ws.cell(i, cp).value
                # 数量规格
                cp_num = extract_num(cp_str)
                # 数量
                # sl_num = ws.cell(i, sl + 2).value
                print(cp_str, cp_num, sl_num)
                # 判断后面连续的是否也为“安瓿”
                j = 1
                next_wl_data = ws.cell(i + 1, wl + 1).value
                while re.search(reg, str(next_wl_data)):
                    i += 1
                    # 数量相加
                    sl_num += ws.cell(i, sl + 2).value
                    next_wl_data = ws.cell(i + j, wl + 1).value
                    j += 1
                # 下面判断数据是否符合要求，误差在0.5内
                if float(sl_num) < float(cp_num) or float(sl_num)-float(cp_num) > 0.5:
                    # 处理
                    print(cp_str, cp_num, sl_num, "数据不合格", j)
                    # 将数据不合格这几行数据标红
                    for r in range(j):
                        print("这条不合格！！！！", r, i-j+r+1)
                        red_fill = PatternFill(fill_type='solid', fgColor="FF0000", bgColor="AACF91")
                        ws.row_dimensions[i-j+r+1].fill = red_fill
                        # green_fill = PatternFill(bgColor="AACF91", fill_type="solid")
                        #
                        # ws.cell(row=i-j+r+1, column=wl).fill = green_fill
            # 是托盘
            elif wl_data is not None and re.search(reg_tp, str(wl_data)):
                # handle_tag(ws, sl_num, i, wl, sl, reg_tp, "0000FF")
                # 数量
                # sl_num = ws.cell(i, sl + 2).value
                # 判断后面连续的是否也为“托盘”
                j = 1
                next_wl_data = ws.cell(i + 1, wl + 1).value
                while re.search(reg_tp, str(next_wl_data)):
                    i += 1
                    # 数量相加
                    sl_num += ws.cell(i, sl + 2).value
                    next_wl_data = ws.cell(i + j, wl + 1).value
                    j += 1
                # 下面判断数据是否符合要求，误差在0.5内
                if float(sl_num) < 1 or float(sl_num) > 1.5:
                    # 处理
                    print("托盘", sl_num, "数据不合格", j)
                    # 将数据不合格这几行数据标红
                    for r in range(j):
                        print("这条不合格！！！！", r, i - j + r + 1)
                        # 蓝
                        red_fill = PatternFill(fill_type='solid', fgColor="0000FF")
                        ws.row_dimensions[i - j + r + 1].fill = red_fill
            # 是彩盒
            elif wl_data is not None and re.search(reg_ch, str(wl_data)):
                # 数量
                # sl_num = ws.cell(i, sl + 2).value
                # 判断后面连续的是否也为“彩盒”
                j = 1
                next_wl_data = ws.cell(i + 1, wl + 1).value
                while re.search(reg_ch, str(next_wl_data)):
                    i += 1
                    # 数量相加
                    sl_num += ws.cell(i, sl + 2).value
                    next_wl_data = ws.cell(i + j, wl + 1).value
                    j += 1
                # 下面判断数据是否符合要求，误差在0.5内
                if float(sl_num) < 1 or float(sl_num) > 1.5:
                    # 处理
                    print("彩盒", sl_num, "数据不合格", j)
                    # 将数据不合格这几行数据标红
                    for r in range(j):
                        print("这条不合格！！！！", r, i - j + r + 1)
                        # 绿
                        red_fill = PatternFill(fill_type='solid', fgColor="00FF00")
                        ws.row_dimensions[i - j + r + 1].fill = red_fill
            # 是说明书
            elif wl_data is not None and re.search(reg_sms, str(wl_data)):
                # 数量
                # sl_num = ws.cell(i, sl + 2).value
                # 判断后面连续的是否也为“彩盒”
                j = 1
                next_wl_data = ws.cell(i + 1, wl + 1).value
                while re.search(reg_sms, str(next_wl_data)):
                    i += 1
                    # 数量相加
                    sl_num += ws.cell(i, sl + 2).value
                    next_wl_data = ws.cell(i + j, wl + 1).value
                    j += 1
                # 下面判断数据是否符合要求，误差在0.5内
                if float(sl_num) < 1 or float(sl_num) > 1.5:
                    # 处理
                    print("说明书", sl_num, "数据不合格", j)
                    # 将数据不合格这几行数据标红
                    for r in range(j):
                        print("这条不合格！！！！", r, i - j + r + 1)
                        # 黑
                        red_fill = PatternFill(fill_type='solid', fgColor="000000")
                        ws.row_dimensions[i - j + r + 1].fill = red_fill
            i += 1
    wb.save("表.xlsx")
    wb.close()






