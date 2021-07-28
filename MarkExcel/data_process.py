import re


# 提取字符串中指定数字
from openpyxl.styles import PatternFill


def extract_num(string):
    reg = r"\*[0-9]+支"
    m = re.search(reg, string)
    # print(m)
    if m is None:
        reg = r":[0-9]+支"
        m = re.search(reg, string)
    print(m)
    # print(m.group()[1:-1])
    return m.group()[1:-1]


'''
处理托盘、说明书、彩盒这3种数据
ws: sheet表
sl_num: （托盘、说明书、彩盒）数量
i: 遍历到达的行数
wl_col: 物料所在列数
sl_col: 数量所在列数
reg: 对应物料名字的正则
color: 将要标记的颜色
'''


def handle_tag(ws, sl_num, i, wl_col, sl_col, reg, color):
    # 数量
    # sl_num = ws.cell(i, sl + 2).value
    # 判断后面连续的是否也为“托盘”、“彩盒”、“说明，说明书”
    j = 1
    next_wl_data = ws.cell(i + 1, wl_col).value
    while re.search(reg, str(next_wl_data)):
        i += 1
        # 数量相加
        sl_num += ws.cell(i, sl_col).value
        next_wl_data = ws.cell(i + j, wl_col).value
        j += 1
    # 下面判断数据是否符合要求，误差在0.5内
    if float(sl_num) < 1 or float(sl_num) > 1.5:
        # 处理
        print(reg, sl_num, "数据不合格", j)
        # 将数据不合格这几行数据标红
        for r in range(j):
            print("这条不合格！！！！", r, i - j + r + 1)
            # 颜色分配
            red_fill = PatternFill(fill_type='solid', fgColor=color)
            ws.row_dimensions[i - j + r + 1].fill = red_fill
    return i


if __name__ == '__main__':
    print(extract_num("10ml:90mg*5支"))
    print(extract_num("10ml:90mg*10支"))



