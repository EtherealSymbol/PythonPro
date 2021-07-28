import re
import openpyxl
from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill

if __name__ == '__main__':

    wb = openpyxl.load_workbook("产品成本计算单查询.xlsx", data_only=True)
    ws = wb.active
    # 获得所有列
    # for col in ws.columns:
    #     print(col)
    rows = ws.rows
    # 获得第一行
    # print(list(rows))
    list_rows = list(rows)
    first_row = list_rows[0]

    # 初始化物料，产品规格，数量为-1
    wl = -1
    cp = -1
    sl = -1
    k = 1
    no = -1
    for elem in first_row:
        # 获得“物料”所在列
        if elem.value == "物料" and wl == -1:
            wl = k
        elif elem.value == "产品规格" and cp == -1:
            cp = k
        elif elem.value == "数量" and sl == -1:
            sl = k
        elif elem.value == "产品编号" and no == -1:
            no = k
        k += 1
    print(wl, cp, sl, no)
    next_rows = list_rows[1:]
    # 正则
    reg = r"安瓿"
    reg_ch = r"彩盒"
    reg_sms = r"说明书"
    reg_tp = r"托盘"
    i = 2

    # 字典
    dict_an = {}
    dict_an_gg = {}
    dict_ch = {}
    dict_sms = {}
    dict_tp = {}

    for row in next_rows:
        for cell in row:
            wl_data = ws.cell(i, wl).value
            # 数量
            sl_num = ws.cell(i, sl).value
            # 产品编号
            cp_no = ws.cell(i, no).value
            # 是安瓿
            if wl_data is not None and re.search(reg, str(wl_data)):
                # 产品规格字符串
                cp_str = ws.cell(i, cp).value
                # 数量规格
                print(cp_str)
                cp_num = extract_num(cp_str)
                # 数量
                # sl_num = ws.cell(i, sl).value
                print(cp_str, cp_num, sl_num, cp_no)
                if dict_an.get(cp_no) is None:
                    dict_an_gg[cp_no] = cp_num
                    dict_an[cp_no] = sl_num
                else:
                    dict_an[cp_no] += sl_num
            # 是托盘
            elif wl_data is not None and re.search(reg_tp, str(wl_data)):
                if dict_tp.get(cp_no) is None:
                    dict_tp[cp_no] = sl_num
                else:
                    dict_tp[cp_no] += sl_num
            # 是彩盒
            elif wl_data is not None and re.search(reg_ch, str(wl_data)):
                if dict_ch.get(cp_no) is None:
                    dict_ch[cp_no] = sl_num
                else:
                    dict_ch[cp_no] += sl_num
    #             # 绿色
    #             i = handle_tag(ws, sl_num, i, wl, sl, reg_ch, "00FF00")
            # 是说明书
            elif wl_data is not None and re.search(reg_sms, str(wl_data)):
                if dict_sms.get(cp_no) is None:
                    dict_sms[cp_no] = sl_num
                else:
                    dict_sms[cp_no] += sl_num
    #             # 黑色
    #             i = handle_tag(ws, sl_num, i, wl, sl, reg_sms, "272727")
            i += 1
    bu_dict_an = {}
    bu_dict_ch = {}
    bu_dict_tp = {}
    bu_dict_sms = {}

    # 判断安瓿
    for key in dict_an:
        print(dict_an[key], dict_an_gg[key])
        if dict_an[key] < float(dict_an_gg[key]) or dict_an[key] - float(dict_an_gg[key]) > 0.5:
            bu_dict_an[key] = 1
    # 判断彩盒
    for key in dict_ch:
        if dict_ch[key] < 1 or dict_ch[key] > 1.5:
            bu_dict_ch[key] = 1
    # 判断托盘
    for key in dict_tp:
        if dict_tp[key] < 1 or dict_tp[key] > 1.5:
            bu_dict_tp[key] = 1

    # 判断说明书
    for key in dict_sms:
        if dict_sms[key] < 1 or dict_sms[key] > 1.5:
            bu_dict_sms[key] = 1

    i = 1
    for row in next_rows:
        for cell in row:
            wl_data = ws.cell(i, wl).value
            # 数量
            sl_num = ws.cell(i, sl).value
            # 产品编号
            cp_no = ws.cell(i, no).value
            # 是安瓿
            if wl_data is not None and re.search(reg, str(wl_data)):
                if bu_dict_an.get(cp_no) is not None:
                    red_fill = PatternFill(fill_type='solid', fgColor="FF0000", bgColor="AACF91")
                    ws.row_dimensions[i].fill = red_fill
            # 是托盘
            elif wl_data is not None and re.search(reg_tp, str(wl_data)):
                if bu_dict_tp.get(cp_no) is not None:
                    red_fill = PatternFill(fill_type='solid', fgColor="0000FF")
                    ws.row_dimensions[i].fill = red_fill
            # 是彩盒
            elif wl_data is not None and re.search(reg_ch, str(wl_data)):
                if bu_dict_ch.get(cp_no) is not None:
                    red_fill = PatternFill(fill_type='solid', fgColor="00FF00")
                    ws.row_dimensions[i].fill = red_fill
            # 是说明书
            elif wl_data is not None and re.search(reg_sms, str(wl_data)):
                if bu_dict_ch.get(cp_no) is not None:
                    red_fill = PatternFill(fill_type='solid', fgColor="272727")
            i += 1
    wb.save("产品成本计算单查询.xlsx")
    wb.close()






