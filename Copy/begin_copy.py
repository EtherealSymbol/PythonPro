import openpyxl
from before_copy import extract_second
from extract_data_to_dict_2021 import extract_data_to_dict_fun
from Cost_in2019.extract_data_to_dict_20191 import extract_20191
from Cost_in2019.extract_data_to_dict_20192 import extract_20192
from Cost_in2019.extract_data_to_dict_20193 import extract_20193
from Cost_in2019.extract_data_to_dict_20194 import extract_20194
from Cost_in2019.extract_data_to_dict_20195 import extract_20195
from Cost_in2019.extract_data_to_dict_20196 import extract_20196
from Cost_in2019.extract_data_to_dict_20197 import extract_20197
from Cost_in2019.extract_data_to_dict_20198 import extract_20198
from Cost_in2019.extract_data_to_dict_20199 import extract_20199
from Cost_in2019.extract_data_to_dict_201910 import extract_201910
from Cost_in2019.extract_data_to_dict_201911 import extract_201911
from Cost_in2019.extract_data_to_dict_201912 import extract_201912


from Cost_in2020.extract_data_to_dict_20201 import extract_20201
from Cost_in2020.extract_data_to_dict_20202 import extract_20202
from Cost_in2020.extract_data_to_dict_20203 import extract_20203
from Cost_in2020.extract_data_to_dict_20204 import extract_20204
from Cost_in2020.extract_data_to_dict_20205 import extract_20205
from Cost_in2020.extract_data_to_dict_20206 import extract_20206
from Cost_in2020.extract_data_to_dict_20207 import extract_20207
from Cost_in2020.extract_data_to_dict_20208 import extract_20208
from Cost_in2020.extract_data_to_dict_20209 import extract_20209
from Cost_in2020.extract_data_to_dict_202010 import extract_202010
from Cost_in2020.extract_data_to_dict_202011 import extract_202011
from Cost_in2020.extract_data_to_dict_202012 import extract_202012

from Cost_in2021.extract_data_to_dict_20211 import extract_20211
from Cost_in2021.extract_data_to_dict_20212 import extract_20212
from Cost_in2021.extract_data_to_dict_20213 import extract_20213
from Cost_in2021.extract_data_to_dict_20214 import extract_20214
from Cost_in2021.extract_data_to_dict_20215 import extract_20215
from Cost_in2021.extract_data_to_dict_20216 import extract_20216

if __name__ == '__main__':
    # file_name = "产品成本计算单查询 - 副本.xlsx"
    file_name = "产品成本计算单查询 - 复制后.xlsx"
    wb = openpyxl.load_workbook(file_name, data_only=True)

    ws = wb['Sheet']

    rows = ws.rows
    list_rows = list(rows)
    # for li in list_rows:
    #     for n in li:
    #         print(n.value, end='  ')
    #     print()
    first_row = list_rows[0]
    # for f in first_row:
    #     print(f.value)

    first_row_len = len(first_row)
    # 逆序遍历，从后往前找物料等信息，因为他们有两列
    # 分别记录他们在第几列，并将其初始化为-1
    cp_no = -1  # 产品编号
    f_wl_no = -1  # 物料编号
    f_wl = -1
    f_gg = -1
    f_dw = -1
    f_tr_sl = -1
    f_tr_je = -1
    f_sl = -1
    f_dj = -1
    f_je = -1
    for i in range(1, first_row_len):
        if first_row[i - 1].value == '产品编号' and cp_no == -1:
            cp_no = i
        if first_row[i - 1].value == '物料编号' and f_wl_no == -1:
            f_wl_no = i
        elif first_row[i - 1].value == '物料' and f_wl == -1:
            f_wl = i
        elif first_row[i - 1].value == '物料规格' and f_gg == -1:
            f_gg = i
        elif first_row[i - 1].value == '物料计量单位' and f_dw == -1:
            f_dw = i
        elif first_row[i - 1].value == '投入数量' and f_tr_sl == -1:
            f_tr_sl = i
        elif first_row[i - 1].value == '投入金额' and f_tr_je == -1:
            f_tr_je = i
        elif first_row[i - 1].value == '数量' and f_sl == -1:
            f_sl = i
        elif first_row[i - 1].value == '单价' and f_dj == -1:
            f_dj = i
        elif first_row[i - 1].value == '金额' and f_je == -1:
            f_je = i
    # print(f_wl_no, f_wl, f_gg, f_dw, f_tr_sl, f_tr_je, f_sl, f_dj, f_je)
    dict_sec = extract_second()

    # dict_compare = extract_data_to_dict_fun("2021年07月小针成本指标表.xls", "2021.07与预算")
    # dict_20211 = extract_20211()
    # dict_20212 = extract_20212()
    # dict_20213 = extract_20213()
    # dict_20214 = extract_20214()
    # 2021年5月
    # dict_20215 = extract_20215()
    # 5月弄过了
    # print(dict_20215['101000326'])
    # dict_20216 = extract_20216()

    # dict_20201 = extract_20201()
    # dict_20202 = extract_20202()
    # dict_20203 = extract_20203()
    # dict_20204 = extract_20204()
    # dict_20205 = extract_20205()
    # dict_20206 = extract_20206()
    # dict_20207 = extract_20207()
    # dict_20208 = extract_20208()
    # dict_20209 = extract_20209()
    # dict_202010 = extract_202010()
    # dict_202011 = extract_202011()
    # dict_202012 = extract_202012()
    # 2019年的
    dict_20191 = extract_20191()
    # dict_20192 = extract_20192()
    # dict_20193 = extract_20193()
    # dict_20194 = extract_20194()
    # dict_20195 = extract_20195()
    # dict_20196 = extract_20196()
    # dict_20197 = extract_20197()
    # dict_20198 = extract_20198()
    # dict_20199 = extract_20199()
    # dict_201910 = extract_201910()
    # dict_201911 = extract_201911()
    # dict_201912 = extract_201912()
    k = 1
    for li in list_rows[1:]:
        cp_no_data = li[cp_no-1].value
        # 第一个物料编号
        f_wl_no_data = li[f_wl_no-1].value
        # 第一个物料
        f_wl_data = li[f_wl-1].value
        # 第二个物料
        s_wl_data = li[dict_sec['s_wl']-1].value
        if f_wl_no_data is not None and f_wl_data is not None and s_wl_data is None:
            # 开始寻找数据并复制数据
            # 这是“产品编号”对应的所有“物料”编号的每月的数据
            # all_month_dy = dict_20211.get(str(cp_no_data))
            # all_month_dy = dict_20212.get(str(cp_no_data))
            # all_month_dy = dict_20213.get(str(cp_no_data))
            # all_month_dy = dict_20214.get(str(cp_no_data))
            # all_month_dy = dict_20215.get(str(cp_no_data))
            # all_month_dy = dict_20216.get(str(cp_no_data))
            # 2020年的
            # all_month_dy = dict_20201.get(str(cp_no_data))
            # all_month_dy = dict_20202.get(str(cp_no_data))
            # all_month_dy = dict_20203.get(str(cp_no_data))
            # all_month_dy = dict_20204.get(str(cp_no_data))
            # all_month_dy = dict_20205.get(str(cp_no_data))
            # all_month_dy = dict_20206.get(str(cp_no_data))
            # all_month_dy = dict_20207.get(str(cp_no_data))
            # all_month_dy = dict_20208.get(str(cp_no_data))
            # all_month_dy = dict_20209.get(str(cp_no_data))
            # all_month_dy = dict_202010.get(str(cp_no_data))
            # all_month_dy = dict_202011.get(str(cp_no_data))
            # all_month_dy = dict_202012.get(str(cp_no_data))
            # 2019年的
            all_month_dy = dict_20191.get(str(cp_no_data))
            # all_month_dy = dict_20192.get(str(cp_no_data))
            # all_month_dy = dict_20193.get(str(cp_no_data))
            # all_month_dy = dict_20194.get(str(cp_no_data))
            # all_month_dy = dict_20195.get(str(cp_no_data))
            # all_month_dy = dict_20196.get(str(cp_no_data))
            # all_month_dy = dict_20197.get(str(cp_no_data))
            # all_month_dy = dict_20198.get(str(cp_no_data))
            # all_month_dy = dict_20199.get(str(cp_no_data))
            # all_month_dy = dict_201910.get(str(cp_no_data))
            # all_month_dy = dict_201911.get(str(cp_no_data))
            # all_month_dy = dict_201912.get(str(cp_no_data))

            if all_month_dy is not None and all_month_dy.get(str(f_wl_no_data)) is not None:
                month_dy = all_month_dy.get(str(f_wl_no_data))
                # 打印数据
                print(cp_no_data, f_wl_no_data, f_wl_data, s_wl_data, month_dy)
                # 开始存数据
                # 物料
                ws.cell(row=k + 1, column=dict_sec['s_wl'], value=month_dy['物料'])
                # 物料规格
                ws.cell(row=k + 1, column=dict_sec['s_gg'], value=month_dy['物料规格'])
                # 物料计量单位
                ws.cell(row=k + 1, column=dict_sec['s_dw'], value=month_dy['物料计量单位'])
                # 物料计量单位
                ws.cell(row=k + 1, column=dict_sec['s_tr_sl'], value=month_dy['投入数量'])
                # 物料计量单位
                ws.cell(row=k + 1, column=dict_sec['s_tr_je'], value=month_dy['投入金额'])
                # 物料计量单位
                ws.cell(row=k + 1, column=dict_sec['s_sl'], value=month_dy['数量'])
                # 物料计量单位
                ws.cell(row=k + 1, column=dict_sec['s_dj'], value=month_dy['单价'])
                # 物料计量单位
                ws.cell(row=k + 1, column=dict_sec['s_je'], value=month_dy['金额'])
                # break
        k += 1
    print("成功！")
    wb.save("产品成本计算单查询 - 复制后.xlsx")
    wb.close()





