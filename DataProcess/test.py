import re
import openpyxl
# from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill

if __name__ == '__main__':
    wb = openpyxl.load_workbook("表.xlsx", data_only=True)
    ws = wb.active
    # 获得所有列
    # for col in ws.columns:
    #     print(col)
    rows = ws.rows
    # 获得第一行
    # print(list(rows))
    list_rows = list(rows)
    first_row = list_rows[0]

    # 初始化物料，产品规格，数量为-1
    wl = -1
    cp = -1
    sl = -1
    k = 1
    for elem in first_row:
        # 获得“物料”所在列
        if elem.value == "物料" and wl == -1:
            wl = k
        elif elem.value == "产品规格" and cp == -1:
            cp = k
        elif elem.value == "数量" and sl == -1:
            sl = k
        else:
            k += 1
    # print(wl, cp, sl)
    next_rows = list_rows[1:]
    # 正则
    reg = r"安瓿"
    reg_ch = r"彩盒"
    reg_sms = r"说明书"
    reg_tp = r"托盘"
    i = 2
    for row in next_rows:
        for cell in row:
            wl_data = ws.cell(i, wl+1).value
            # 数量
            sl_num = ws.cell(i, sl + 2).value
            # 产品规格字符串
            cp_str = ws.cell(i, cp).value
            # 数量规格
            # cp_num = extract_num(cp_str)
            # 数量
            # sl_num = ws.cell(i, sl + 2).value
            print(wl_data, sl_num, cp_str)


