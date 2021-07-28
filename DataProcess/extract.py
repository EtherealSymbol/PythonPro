import openpyxl
from basic_op import MyReadExcel

# if __name__ == '__main__':
    # filename = '表2.xlsx'
    # # 读取数据
    # wb = openpyxl.load_workbook(filename, data_only=True)
    # sheets = wb.sheetnames  # 获取sheet页
    # new_wb = openpyxl.Workbook()
    # row = ["物料编码", "数量（盒）"]
    # # 2.向工作表中 按行添加数据
    # new_ws = new_wb["Sheet"]
    # new_ws.append(row)
    # # for i in range(0, len(sheets)):
    # #     sheet = wb[sheets[i]]
    # #     print(sheet)
    # all_datas = []
    # for i in range(0, len(sheets)):
    #     r = MyReadExcel(filename, sheets[i])
    #     all_datas.append(r.read_data())
    # for all_data in all_datas:
    #     for data in all_data:
    #         # print(data)
    #         if data[3] is not None and str(data[3]).isdigit():
    #             # print(data[3], data[7])
    #             data_row = [data[3], data[7]]
    #             # 2.向工作表中 按行添加数据
    #             new_ws.append(data_row)
    # new_wb.save("中间表.xlsx")
    # print("创建完成")

'''
将要合并的数据存入字典，方便之后索取
'''


def extract_to_dict(file_name):
    # 读取数据
    wb = openpyxl.load_workbook(file_name, data_only=True)
    sheets = wb.sheetnames  # 获取sheet页
    new_wb = openpyxl.Workbook()
    row = ["物料编码", "数量（盒）"]
    # 2.向工作表中 按行添加数据
    new_ws = new_wb["Sheet"]
    new_ws.append(row)
    # for i in range(0, len(sheets)):
    #     sheet = wb[sheets[i]]
    #     print(sheet)
    all_datas = []
    for i in range(0, len(sheets)):
        r = MyReadExcel(file_name, sheets[i])
        all_datas.append(r.read_data())
    dicts = {}
    for all_data in all_datas:
        for data in all_data:
            # print(data)
            if data[3] is not None and str(data[3]).isdigit():
                dicts[str(data[3])] = data[7]
    print("提取并创建字典完成")
    return dicts

